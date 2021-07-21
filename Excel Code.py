import openpyxl
import pandas as pd

x1 = openpyxl.load_workbook("Rounds.xlsx")
sheet1 = x1["Round 7"]

lista = [str(sheet1.cell(row = a, column = 10).value) for a in range(7, 55)]

results1column = []

for a in lista:
    if a[0] != "½":
        results1column.append(int(a[0]))
    else:
        results1column.append(0.5)

results2column = []

for a in lista:
    if a[-1] != "½":
        results2column.append(int(a[-1]))
    else:
        results2column.append(0.5)

#print(results1column) # results of round 7 for left hand column of players in excel
#print(results2column) # results of round 7 for right hand column of players


lista2 = [sheet1.cell(row = a, column = 9).value for a in range(7, 55)]

points1 = []

for a in lista2:
    if type(a) is int:
        points1.append(a)
    elif a[0] != "½":
        points1.append(int(a[0])+ 0.5)
    else:
        points1.append(0.5)

#print(points1)

lista3 = [sheet1.cell(row = a, column = 11).value for a in range(7, 55)]

points2 = []

for a in lista3:
    if type(a) is int:
        points2.append(a)
    elif a[0] != "½":
        points2.append(int(a[0])+ 0.5)
    else:
        points2.append(0.5)

totalpoints1 = [a + b for (a, b) in zip(results1column, points1)]
totalpoints2 = [a + b for (a, b) in zip(results2column, points2)]



